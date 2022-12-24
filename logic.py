import openpyxl
import PyPDF2
import re
import codecs
import pandas as pd
import chardet
import os
import subprocess


qPDF = "Questões.pdf"
gPDF = "Gabarito.pdf"
gabaritos = []
questoes = []
QUESTregex = r'[0-9]+[\s]+[-][\s]+[2][0][2][2]'
GABregex = r'[0-9]+[\s]+[-][\s]+[2][0][2][2]'


def extract_text_from_pdf(pdf_file):
    # Open the PDF file in read-binary mode
    with open(pdf_file, 'rb') as file:
        # Create a PDF object
        pdf = PyPDF2.PdfFileReader(file)
        # Iterate over every page
        text = ''
        for page in range(pdf.getNumPages()):
            # Extract the text from the page
            page_text = pdf.getPage(page).extractText()
            # Append the text to the overall text
            text += page_text
        # Return the text
        return text

# identify the parts of the text which are questions or gabaritos, depending on the type specified


def create_substrings(text, array):
    # Create an empty list to store the substrings
    parts = []
    # Iterate over the pairs of start and end indices
    for start, end in zip(array[:-1], array[1:]):
        # Append the substring to the list
        parts.append(text[start:end])
    # Add the final substring from the last index to the end of the text
    parts.append(text[array[-1]:])
    return parts


def identify(text, type):
    parts = []
    if type == "gabarito":
        # Use a regular expression to find all occurrences in the text
        matches = list(re.finditer(GABregex, text))
        for match in matches:
            # Get the start and end indices of the match
            gabaritos.append(match.start())
        parts = create_substrings(text, gabaritos)
    if type == "questões" or type == "questoes":
        matches = list(re.finditer(QUESTregex, text))
        for match in matches:
            # Get the start and end indices of the match
            questoes.append(match.start())
        parts = create_substrings(text, questoes)
    return parts


# format each substring based on the type

def format_questoes(questoesArray):
    questoesFormatadas = []
    # Iterate over the questions in the array
    for questao in questoesArray:
        # Remove all the extra whitespace
        questao = re.sub(" +", " ", questao)
        questao = re.sub("\t+", " ", questao)
        # Remove all the breaklines but the first one
        replacement = " "
        # Obtém a primeira ocorrência de \n
        first_newline_index = questao.index('\n')
        # Substitui todas as ocorrências de \n a partir da segunda ocorrência
        questao = questao[:first_newline_index + 1] +\
            questao[first_newline_index:].replace('\n', replacement)
        # Add a breakline before each match of '[A-E][)][ ]'
        questao = re.sub(r'([ ][A-E][)][ ])', r'\n\1', questao)
        questao = re.sub(r'([ ][A-E][)][ ])', r'\n\1', questao)
        # Append the formatted question to the list
        if "Realize as questões de concursos selecionadas" in questao:
            questao = questao[:questao.index(
                'Questões da Apostila')]
        questoesFormatadas.append(questao)
    # Return the list of formatted questions
    return questoesFormatadas


def format(type, array):
    if type == "gabaritos":
        return format_gabaritos(array)
    if type == "questoes":
        return format_questoes(array)


def format_gabaritos(gabaritosArray):
    gabaritosFormatados = []
    # Iterate over the gabaritos in the array
    for gabarito in gabaritosArray:
        gabarito = re.sub(" +", " ", gabarito)
        # Add a breakline before each match of '[A-E][)][ ]'
        gabarito = re.sub("\t+", " ", gabarito)
        # Obtém a primeira ocorrência de \n
        replacement = ' '
        first_newline_index = gabarito.index('\n')
        # Substitui todas as ocorrências de \n a partir da segunda ocorrência
        gabarito = gabarito[:first_newline_index + 1] + \
            gabarito[first_newline_index + 1:].replace('\n', replacement)
        # Remove all the extra whitespace
        # Append the formatted question to the list
        if "Comentários da equipe acadêmica" in gabarito:
            gabarito = gabarito[:gabarito.index(
                'Comentários da equipe acadêmica')]
        gabaritosFormatados.append(gabarito)
    # Return the list of formatted questions
    return gabaritosFormatados


# Get the gabaritos (just the letters)

def get_Gab_Letters(gabaritoPDF):
    # Open the PDF file in read-binary mode
    with open(gabaritoPDF, 'rb') as file:
        # Create a PDF object
        pdf = PyPDF2.PdfFileReader(file)
        # Get first page text
        text = pdf.getPage(0).extractText()
        # Format text to make it easier to match
        text = re.sub('\t+', ' ', text)
        # Find index of "Legenda" which will serve as the end index of our string
        index = text.index('Legenda')
        # Creating new string containing the gabs
        newText = text[0:index - 1]
        # Creating Gabs array
        Gabs = newText.split('\n')
        return Gabs


questoesFormatadas = []
gabaritosFormatados = []


def init(QuestoesPdf, GabaritoPdf, type1, type2):
    Questoestext = extract_text_from_pdf(QuestoesPdf)
    Gabaritotext = extract_text_from_pdf(GabaritoPdf)
    questParts = identify(Questoestext, type1)
    gabaritoParts = identify(Gabaritotext, type2)
    gabaritosFormatados = format("gabaritos", gabaritoParts)
    questoesFormatadas = format("questoes", questParts)
    Gabs = get_Gab_Letters(GabaritoPdf)

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    openpyxl.load_workbook('sample.xlsx')

    # Get the active sheet
    sheet = workbook.active

    # Iterate over the questions in the array
    for i, questao in enumerate(questoesFormatadas):
        # Write the question to the first row and first column
        sheet.cell(row=i+1, column=1).value = questao
    # Iterate over the gabaritos in the array
    for i, gabarito in enumerate(gabaritosFormatados):
        # Write the gabarito to the third row and first column
        sheet.cell(row=i+1, column=3).value = gabarito

    # Iterate over the gabs in the array
    for i, gab in enumerate(Gabs):
        # Write the gabarito to the third row and first column
        sheet.cell(row=i+1, column=2).value = gab

    # Save the workbook to a file
    workbook.save("questoes_e_gabaritos.xlsx")

    # Lê o arquivo .xlsx como um DataFrame do pandas
    df = pd.read_excel('questoes_e_gabaritos.xlsx')

    # Escreve o DataFrame em um arquivo .txt usando o separador de tabulação
    df.to_csv('questoes_e_gabaritos.txt', sep='\t', index=False)

    encoding = ''
    with open('questoes_e_gabaritos.txt', 'rb') as f:
        data = f.read()
        encoding = chardet.detect(data)['encoding']

    # Abre o arquivo original com a codificação atual
    with codecs.open('questoes_e_gabaritos.txt', 'r', encoding=encoding) as f:
        # Lê o conteúdo do arquivo
        text = f.read()

    # Abre o arquivo de saída com a codificação UTF-8
    with codecs.open('questoes_e_gabaritos.txt', 'w', encoding='utf-8') as f:
        # Escreve o conteúdo no arquivo de saída
        f.write(text)

    old_questoes_name = 'Questões.pdf'
    old_gabarito_name = 'Gabarito.pdf'
    new_questoes_name = 'Questões-ANTIGO.pdf'
    new_gabarito_name = 'Gabarito-ANTIGO.pdf'

    os.remove(new_questoes_name)
    os.remove(new_gabarito_name)

    os.rename(old_questoes_name, new_questoes_name)
    os.rename(old_gabarito_name, new_gabarito_name)


init(qPDF, gPDF, "questoes", "gabarito")
